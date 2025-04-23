import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def save_to_excel(blogs, file_name="blogs.xlsx"):
    df = pd.DataFrame([{
        "Title": blog["title"],
        "URL": blog["url"],
        "Content": blog["content"][:500] + ("..." if len(blog["content"]) > 500 else ""),
        "VADER Compound": blog["vader_compound"],
        "VADER Sentiment": blog["vader_sentiment"],
        "Empath Tags": blog["empath_tags"],
    } for blog in blogs])

    df.to_excel(file_name, index=False)
    wb = load_workbook(file_name)
    ws = wb.active

    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ws.columns:
        length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 5, 50)

    wb.save(file_name)
